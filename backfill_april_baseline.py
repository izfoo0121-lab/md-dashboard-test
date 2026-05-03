#!/usr/bin/env python3
"""
Backfill April 2026 patronage opening baselines from the 2026-04-30 debtor file.

Requires:
  SUPABASE_URL
  SUPABASE_SERVICE_KEY
"""

import json
import os
import sys
import urllib.parse
import urllib.request
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


MONTH = "2026-04"
SNAPSHOT_DATE = "2026-04-01"
SOURCE_FILE = Path(r"C:\Users\tgy_3\Downloads\debtor maintainance3004 (1).xlsx")
NOTES = (
    "BACKFILL: reconstructed from 2026-04-30 maintenance file "
    "(today_count minus April onboards). Approximate; not a true Apr 1 snapshot."
)

EXPECTED = {
    "BEN": 90,
    "CJ": 102,
    "JACKY": 142,
    "JAMES": 125,
    "KEAN": 115,
    "KEE": 127,
    "KF": 133,
    "KI-MI": 155,
    "KW": 137,
    "LEON": 97,
    "NMK": 115,
    "SAM": 48,
    "YI": 167,
}


def _norm(value):
    return str(value).strip() if value is not None else ""


def _parse_date(value):
    text = _norm(value)
    if not text or text.lower() in ("nan", "none"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    except Exception:
        return None


def compute_counts(path=SOURCE_FILE):
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    required = ["Code", "Debtor Type", "Agent", "Active", "Open Acct Date"]
    missing = [h for h in required if h not in idx]
    if missing:
        raise RuntimeError(f"Missing required columns: {missing}")

    cutoff = date(2026, 4, 1)
    counts = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = _norm(row[idx["Code"]])
        if not code:
            continue
        active = _norm(row[idx["Active"]]).lower()
        if active != "checked":
            continue
        debtor_type = _norm(row[idx["Debtor Type"]])
        if not debtor_type or debtor_type.lower() in ("nan", "none"):
            continue
        if debtor_type.upper() in ("P-PERSONAL", "PERSONAL"):
            continue
        agent = _norm(row[idx["Agent"]]).upper()
        if not agent:
            continue
        open_date = _parse_date(row[idx["Open Acct Date"]])
        if open_date is not None and open_date >= cutoff:
            continue
        counts[agent] += 1
    return dict(sorted(counts.items()))


def _supabase_request(path, method="GET", body=None):
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in the environment")
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(
        f"{url.rstrip('/')}/rest/v1/{path}",
        data=data,
        method=method,
        headers={
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Prefer": "resolution=merge-duplicates,return=representation",
        },
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        text = resp.read().decode("utf-8")
        return json.loads(text) if text else None


def print_counts(counts):
    for agent in EXPECTED:
        print(f"{agent}: {counts.get(agent, 0)}")
    print(f"TOTAL: {sum(counts.values())}")


def validate_counts(counts):
    if counts != EXPECTED:
        print("Computed counts do not match accepted baseline values.", file=sys.stderr)
        print("Computed:", counts, file=sys.stderr)
        print("Expected:", EXPECTED, file=sys.stderr)
        return False
    return True


def upsert_counts(counts):
    existing = _supabase_request(
        f"patronage_history?select=agent&month=eq.{urllib.parse.quote(MONTH)}"
    ) or []
    existing_agents = {r.get("agent") for r in existing}
    rows = [
        {
            "month": MONTH,
            "agent": agent,
            "opening_total": counts[agent],
            "snapshot_date": SNAPSHOT_DATE,
            "notes": NOTES,
        }
        for agent in EXPECTED
    ]
    returned = _supabase_request(
        "patronage_history?on_conflict=month,agent",
        method="POST",
        body=rows,
    ) or []
    print("\nUpsert summary:")
    for row in returned:
        agent = row["agent"]
        status = "updated" if agent in existing_agents else "inserted"
        print(f"{agent:6} {row['opening_total']:4} {status}")
    print(f"\n{len(returned)} rows upserted to patronage_history for month {MONTH}")


def main():
    counts = compute_counts()
    print_counts(counts)
    if not validate_counts(counts):
        return 1
    upsert_counts(counts)
    print("\nVerification SQL:")
    print(
        "SELECT month, agent, opening_total, snapshot_date, notes\n"
        "FROM patronage_history\n"
        "WHERE month = '2026-04'\n"
        "ORDER BY agent;"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
